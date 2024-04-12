import openpyxl as ox
import asyncio

from pathlib import Path
from typing import Union
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from aiogram.methods.edit_message_text import EditMessageText
from aiogram.types import (InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery,
                           InputMediaPhoto)

from sql.models.categories_models import (PhonesTable, TvTable, LaptopsTable, CosmeticsTable,
                                          PowerToolsTable, ElectronicTable, ChildrenGoodsTable)
from sql.sql_engine import SQLEngine
from sql.models import Base, AllProductsTable

from keyboards.catalog_kb import CategoryPage, ProductPage, Photo, PhotoAction

TableType = Union[PhonesTable, TvTable, LaptopsTable, CosmeticsTable, PowerToolsTable,
ElectronicTable, ChildrenGoodsTable]

table_matches = {
    PhonesTable: ["Смартфоны", PhonesTable],
    TvTable: ["Телевизоры", TvTable],
    LaptopsTable: ["Ноутбуки", LaptopsTable],
    CosmeticsTable: ["Лэтуаль", CosmeticsTable],
    PowerToolsTable: ["Электроинструменты", PowerToolsTable],
    ElectronicTable: ["Электроника", ElectronicTable],
    ChildrenGoodsTable: ["Детские товары", ChildrenGoodsTable]
}


class ProductPageKeyboard:
    def __init__(self,
                 catalog_callback_factory: CategoryPage | str = "",
                 product_callback_factory: ProductPage | str = ""):
        self.catalog_callback_factory = catalog_callback_factory
        self.product_callback_factory = product_callback_factory

    def __keyboard_structure(self,
                             user_type,
                             number_of_photos,
                             current_photo_number):

        if user_type == "user":
            default_kb = [["Добавить в корзину 🛒"],
                          ["Назад в каталог 📲"]]

            if 0 <= number_of_photos <= 1:
                return default_kb
            elif number_of_photos > 1 and current_photo_number == 0:
                default_kb.insert(0, ["Следующее фото 📸"])
                return default_kb
            elif number_of_photos == current_photo_number + 1:
                default_kb.insert(0, ["Предыдущее фото 📸"])
                return default_kb
            else:
                default_kb.insert(0, ["⏪", f"{current_photo_number + 1}/5", "⏩"])
                return default_kb

        elif user_type == "admin":
            default_kb = [["Назад к выбору товара"]]

            if number_of_photos < 5:
                default_kb.insert(0, ["Добавить фото"])

            if number_of_photos > 0 and current_photo_number == 0:
                default_kb.insert(0, ["Следующее фото"])
                default_kb.insert(2, ["Удалить фото"])
                return default_kb
            elif current_photo_number + 1 == number_of_photos:
                default_kb.insert(0, "Предыдущее фото")
                return default_kb
            elif 1 < current_photo_number + 1 < 5:
                default_kb.insert(0, ["⏪", f"{current_photo_number + 1}/5", "⏩"])
                return default_kb

            return default_kb

    def set_back_button(self):
        if isinstance(self.product_callback_factory, ProductPage):
            number_of_photos = self.product_callback_factory.number_of_photos
            self.catalog_callback_factory.return_after_photo = True if number_of_photos else False
            self.catalog_callback_factory.paginator = False
        else:
            print("self.product_callback_factory не определён")
            return

    def __keyboard_structure_iteration(self, default_kb: list[list[str]]):

        inline_kb: list[list[InlineKeyboardButton]] = []

        inline_button = lambda text, callback_data: InlineKeyboardButton(
            text=text,
            callback_data=callback_data)

        cf_model = lambda: self.product_callback_factory.model_copy()

        for button_row in default_kb:
            button_row_list = []
            for button_text in button_row:

                if button_text in ["Следующее фото 📸", "⏩", "Следующее фото"]:
                    cf = cf_model()
                    cf.current_photo += 1
                    button_row_list.append(inline_button(button_text, cf.pack()))

                if button_text in ["Предыдущее фото 📸", "⏪", "Предыдущее фото"]:
                    cf = cf_model()
                    cf.current_photo -= 1
                    button_row_list.append(inline_button(button_text, cf.pack()))

                if button_text == "Добавить в корзину 🛒":
                    button_row_list.append(inline_button(button_text, "cart"))

                if button_text in ["Назад в каталог 📲", "Назад к выбору товара"]:
                    self.set_back_button()
                    button_row_list.append(inline_button(
                        button_text, self.catalog_callback_factory.pack()))
                inline_kb.append(button_row_list)

                if button_text == "Добавить фото":
                    button_row_list.append(inline_button(
                        button_text,
                        Photo(photo_action=PhotoAction.add_photo).pack()))

                if button_text == "Удалить фото":
                    button_row_list.append(inline_button(button_text, "remove_photo"))

                if button_text[-2:] == "/5":
                    button_row_list.append(inline_button(button_text, "count_photo"))

        return InlineKeyboardMarkup(inline_keyboard=inline_kb)

    def build_keyboard(self):
        if isinstance(self.product_callback_factory, ProductPage):
            cf = self.product_callback_factory

            user_type = cf.user_type
            number_of_photos = cf.number_of_photos
            current_photo_number = cf.current_photo

            default_kb = self.__keyboard_structure(user_type,
                                                   number_of_photos,
                                                   current_photo_number)

            reply_kb = self.__keyboard_structure_iteration(default_kb)

            return reply_kb


class ProductPageBase:
    categories_dict = {
        "phones": ["Смартфоны", PhonesTable],
        "cosmetic": ["Лэтуаль", CosmeticsTable],
        "children_goods": ["Детские товары", ChildrenGoodsTable],
        "electronic": ["Электроника", ElectronicTable],
        "power_tools": ["Электроинструменты", PowerToolsTable],
        "tvs": ["Телевизоры", TvTable],
        "laptops": ["Ноутбуки", LaptopsTable]
    }

    def __init__(self):
        self.excel_template_path: Path = Path.cwd() / "excel_files/Обновленный Excel-файл.xlsx"
        self.excel_template_path_2: Path = Path.cwd().parents[1] / "excel_files/Обновленный Excel-файл.xlsx"
        self.sql_engine = SQLEngine()

        self.__keyboard: ProductPageKeyboard = ProductPageKeyboard()

        self.category_page_cf: CategoryPage | None = None
        self.product_page_cf: ProductPage | None = None

    def set_keyboard(self, catalog_callback_factory=None, product_callback_factory=None):
        if catalog_callback_factory:
            self.__keyboard.catalog_callback_factory = catalog_callback_factory

        if product_callback_factory:
            self.__keyboard.product_callback_factory = product_callback_factory

        return

    def get_keyboard(self):
        return self.__keyboard.build_keyboard()

    def get_name_and_model(self, category_name: str) -> tuple:
        name, model = self.categories_dict.get(category_name)
        return name, model

    def get_columns_names_from_excel(self):
        path = self.excel_template_path_2
        wb = ox.load_workbook(filename=path)

        # Достаем названия категорий из файла excel
        sheetnames = [sheetname for sheetname in wb.sheetnames
                      if sheetname not in ["Заявки", "На продажу"]]

        return sheetnames

    async def get_attrs_of_a_spec_category(self, table: TableType):
        sheetnames = self.get_columns_names_from_excel()

        category_name = table_matches[table]

        await self.sql_engine.display_a_product_of_a_specific_category(category=category_name)

        return category_name

    def get_column_names_of_category(self, category_name):
        name, model = self.get_name_and_model(category_name)

        excel_column_names = []

        wb = ox.load_workbook(self.excel_template_path)
        ws = wb[name]

        for column_number in range(3, ws.max_column + 1):
            excel_column_names.append(ws.cell(row=1, column=column_number).value)

        return excel_column_names, model

    async def get_product_info(self, category_name: str,
                               session: AsyncSession, product_id: int):

        excel_column_names, model = self.get_column_names_of_category(category_name=category_name)

        model: Base
        excel_column_names: list

        catalog_product = await session.get(AllProductsTable, product_id)
        product_name, product_price = catalog_product.name, catalog_product.price

        stmt = select(model).where(model.product_id == product_id)

        res = await session.execute(stmt)
        result = res.scalars().one()

        column_names = model.collect_column_names()

        product_data = []

        for column_name in column_names:
            if column_name == "product_id":
                continue
            product_data.append(result.__getattribute__(column_name))

        product_information = (
            f"<b><i>{product_name}</i></b>\n\n"
            f"<b><i>{'{:,}'.format(product_price).replace(",", " ")} руб</i></b>\n\n")

        for excel_cols_name, curr_data in zip(excel_column_names, product_data):

            font_excel_cols_name = f"<b><i>{excel_cols_name}</i></b>"
            font_curr_data = f"<i>{curr_data}</i>"

            if curr_data == "None":
                continue
            elif excel_cols_name == "Дополнительное описание":
                product_information += f"\n<b>{excel_cols_name}</b>\n{curr_data}"
            else:
                product_information += f"{font_excel_cols_name}: {font_curr_data}\n"

        return product_information

    async def product_page_answer(self,
                                  callback: CallbackQuery,
                                  text: str,
                                  photos: list = None,
                                  previous_message_id: int = None,
                                  chat_id: int = None,
                                  first_press: bool = False):

        cf = self.__keyboard.product_callback_factory
        cf2 = self.__keyboard.catalog_callback_factory

        number_of_photos = cf.number_of_photos
        current_photo = cf.current_photo

        reply_kb = self.get_keyboard()

        edit_message = callback.message.edit_text(text, reply_markup=reply_kb)

        if number_of_photos == 0:
            await edit_message

        else:

            photo = photos[current_photo]

            delete_bot_message = callback.bot.delete_message(
                chat_id=chat_id,
                message_id=previous_message_id)

            delete_message = callback.message.delete()

            answer_photo = callback.message.answer_photo(
                photo=photo,
                caption=text,
                reply_markup=reply_kb
            )

            edit_photo_message = callback.message.edit_media(
                media=InputMediaPhoto(media=photo, caption=text),
                reply_markup=reply_kb)

            if first_press:
                await delete_bot_message if previous_message_id else await delete_message
                await answer_photo
            else:
                await edit_photo_message

    def photo_addition_process(self,
                               callback: CallbackQuery,
                               text: str) -> tuple[EditMessageText, int]:

        return_callback = self.__keyboard.product_callback_factory.pack()

        reply_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Отмена", callback_data=return_callback)]
        ])

        message_answer = callback.message.edit_text(
                            text=text,
                            reply_markup=reply_keyboard)

        return message_answer, message_answer.message_id





