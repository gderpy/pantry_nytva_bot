import logging
import os

import openpyxl as ox

from pathlib import Path
from openpyxl import Workbook

from aiogram.types import BufferedInputFile, Message
from aiogram import Bot
from sqlalchemy import select, Integer
from sqlalchemy.orm import DeclarativeBase

from sql.sql_engine import SQLEngine
from sql.models.categories_models import (PhonesTable, TvTable, LaptopsTable, CosmeticsTable,
                                          PowerToolsTable, ElectronicTable, ChildrenGoodsTable)
from sql.models import Base, SellsTable, OrdersTable, AllProductsTable


class ExcelParser:

    bot: Bot
    chat_id: int

    def __init__(self):
        self.excel_template_path: Path = Path.cwd() / "excel_files/Обновленный Excel-файл.xlsx"
        self.session = SQLEngine().async_session
        self.sql_engine = SQLEngine()
        self.sheet_names = ["Смартфоны", "Лэтуаль", "Детские товары", "Электроника",
                            "Электроинструменты", "Телевизоры", "Ноутбуки"]

    def excel_file(self):
        return BufferedInputFile(open(self.excel_template_path, "rb").read(),
                                 filename="Ваш Excel-файл.xlsx")

    @staticmethod
    def __unload_orders_and_sells(sell_table_data: list,
                                  order_table_data: list,
                                  wb: Workbook):

        obj_data_list = [sell_table_data, order_table_data]

        for obj_data in obj_data_list:
            for ir in range(len(obj_data)):
                for ic in range(len(obj_data[ir])):

                    if obj_data == sell_table_data:
                        sheet_name = "На продажу"
                        startrow, startcol = 3, 1
                    else:
                        startrow, startcol = 3, 1
                        sheet_name = "Заявки"

                    wb[sheet_name].cell(startrow + ir, startcol + ic).value = obj_data[ir][ic]

    def __define_model_for_sheet_name(self, sheet_name: str):
        if sheet_name not in self.sheet_names:
            raise ValueError("Данная категория отсутствует в БД")

        if sheet_name == "Смартфоны":
            return PhonesTable
        elif sheet_name == "Лэтуаль":
            return CosmeticsTable
        elif sheet_name == "Детские товары":
            return ChildrenGoodsTable
        elif sheet_name == "Электроника":
            return ElectronicTable
        elif sheet_name == "Электроинструменты":
            return PowerToolsTable
        elif sheet_name == "Телевизоры":
            return TvTable
        elif sheet_name == "Ноутбуки":
            return LaptopsTable

    @staticmethod
    def __format_the_value_in_the_numeric_columns(value: str):
        only_digits = "".join([x for x in value if x.isdigit()])
        if only_digits:
            return int(only_digits)

    @staticmethod
    def __get_a_list_of_integer_columns(model: DeclarativeBase):

        integer_columns = []

        for column_name in model.collect_column_names():
            column = getattr(model, column_name)
            column_type = column.type

            if isinstance(column_type, Integer):
                integer_columns.append(column_name)

        return integer_columns

    async def __read_excel_file(self, path):
        path = path
        wb = ox.load_workbook(filename=path)

        excel_data = {}

        await self.sql_engine.clear_the_tables(model=AllProductsTable)

        for sheet_name in self.sheet_names:

            ws = wb[sheet_name]
            excel_data[sheet_name] = {}

            # excel_data = {"Смартфоны": {}}

            model = self.__define_model_for_sheet_name(sheet_name=sheet_name)
            integer_columns = self.__get_a_list_of_integer_columns(model)

            logging.info(f"integer_columns: {integer_columns}")

            for row_number in range(2, ws.max_row + 1):

                if ws.cell(row=row_number, column=1).value is not None:
                    excel_data[sheet_name][row_number - 1] = {}
                    # excel_data = {"Смартфоны": {1: {}}}
                    row_data = []

                    # Получаем название и стоимость товара
                    get_name_price = lambda x: ws.cell(row=row_number, column=x).value
                    product = {"name": get_name_price(1),
                               "price": get_name_price(2),
                               "category": sheet_name}

                    # Вставляем данные в нашу таблицу
                    await self.sql_engine.insert_objects(AllProductsTable, data=product)

                    product_id = await self.sql_engine.get_last_id()

                    # Проходимся циклом по колонкам
                    for column_number in range(3, ws.max_column + 1):

                        cell_value = str(ws.cell(row=row_number, column=column_number).value)

                        row_data.append(cell_value)

                        for key, value in zip(model.collect_column_names(), row_data):
                            print(f"key: {key}, integer_columns: {integer_columns}")

                            if key in integer_columns and value is not None:
                                print(value)
                                if isinstance(value, str):
                                    value = self.__format_the_value_in_the_numeric_columns(value)

                            excel_data[sheet_name][row_number - 1][key] = value
                            excel_data[sheet_name][row_number - 1]["product_id"] = product_id
                else:
                    break

        return excel_data

    async def upload_data_from_excel_file_to_models(self, path):
        excel_data = await self.__read_excel_file(path)

        for category in excel_data:
            model = self.__define_model_for_sheet_name(category)

            await self.sql_engine.clear_the_tables(model=model)

            for row in excel_data[category]:
                data = excel_data[category][row]
                await self.sql_engine.insert_objects(model=model, data=data)

    async def __get_data_from_db_table(self, model: Base) -> list:
        async with self.session() as session:
            stmt = select(model)
            res = await session.execute(stmt)

            obj_data: list = []

            for data in res.scalars():
                print(data)
                obj_data.append(data.as_list())

            await session.commit()
            return obj_data

    async def download_excel_file(self):

        obj_data_orders = await self.__get_data_from_db_table(model=OrdersTable)
        obj_data_sells = await self.__get_data_from_db_table(model=SellsTable)

        wb = ox.load_workbook(filename=self.excel_template_path)

        self.__unload_orders_and_sells(order_table_data=obj_data_orders,
                                       sell_table_data=obj_data_sells,
                                       wb=wb)

        wb.save(filename=self.excel_template_path)

    @staticmethod
    async def save_the_attached_file(message: Message, file_id, file_path):
        file = await message.bot.get_file(file_id=file_id)
        downloaded_file = await message.bot.download_file(file.file_path)
        save_dir = Path.cwd() / "excel_files"
        save_path = os.path.join(save_dir, file_path)

        with open(save_path, "wb") as new_file:
            new_file.write(downloaded_file.read())

        return save_path































